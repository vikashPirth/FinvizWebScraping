import requests
from bs4 import BeautifulSoup
import pandas as pd 
from openpyxl import load_workbook
from io import StringIO

workbook = load_workbook('FinvizData.xlsx')
sheet = workbook["Sheet1"]


def findStartPoint():
    index = 1
    #if this loop is true, the current cell has a value inside of it
    while(sheet.cell(row = index, column = 1).value != None):
        index += 1
    return index

def putDataIntoSheet(df):
    rowIndex = findStartPoint() - 1
   
    df = df.iloc[:, 1:]
    for ticker in range(len(df)):
        columnIndex = 1
        rowIndex += 1
        #find the row corresponding to the ticker index
        tickerDetails = df.iloc[ticker]
        for tickerDetail in tickerDetails:
            sheet.cell(row = rowIndex, column=columnIndex).value = tickerDetail
            columnIndex += 1
    

def getTotalPages(soup):

    try:
        paginationTags = soup.find(class_ = "body-table screener_pagination").find_all('a')
    except Exception as e:
        print("Page is empyt")
        return
    totalPages = (len(paginationTags) - 1)
    
    if (totalPages == 0):
        return 1
    else:
        return totalPages


def getWebpage():
    url = "https://finviz.com/screener.ashx?v=121&f=cap_large,fa_div_o5&ft=4"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}
    reponse = requests.get(url=url, headers=headers)
    soup  = BeautifulSoup(reponse.content, 'html.parser')
    
    numPage = getTotalPages(soup=soup)
    tickerNumber = 1
    
    for numPage in range(numPage):
        response = requests.get(url= url+f"&r={tickerNumber}", headers=headers)
        soup = BeautifulSoup(response.content, 'html.parser')
        table = soup.find('table', class_='styled-table-new is-rounded is-tabular-nums w-full screener_table')
        table_html = StringIO(str(table))
        pdData = pd.read_html(table_html)
        tickerNumber+=20
        putDataIntoSheet(pdData[0])

getWebpage()

workbook.save("FinvizData.xlsx")